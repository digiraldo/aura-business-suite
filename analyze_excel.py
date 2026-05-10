import openpyxl
import csv
from datetime import datetime

XLSX_PATH = r'c:\laragon\www\diserwp\wp-content\plugins\aura-business-suite\Y-Global2 2017-20 (Use this one).xlsx'
OUT_CSV   = r'c:\laragon\www\diserwp\wp-content\plugins\aura-business-suite\import_2026.csv'

INCOME_COLS = {
    16: 'Alojamiento Nacionales',
    17: 'Renta Vidal/Schmidt',
    18: 'Alojamiento Internacional',
    19: 'Alojamiento Programas Hadime',
    20: 'Renta Van',
    21: 'Escuela de Lenguaje',
    22: 'Ingresos Hadime sin alojamiento',
    23: 'Intereses y Otros Ingresos',
}
EXPENSE_COLS = {
    24: 'Productos Limpieza',
    25: 'Salario Limpieza',
    26: 'Mantenimiento',
    27: 'Gastos Equipo',
    28: 'Honorarios Contables',
    29: 'Cuidado Perro',
    30: 'Internet',
    31: 'Salario Voluntariado',
    32: 'Alberca e Irrigacion',
    33: 'Comisiones Bancarias',
    34: 'Relaciones Publicas',
    35: 'Gastos Varios Voluntarios',
    36: 'Escuela Lenguaje Egreso',
    37: 'Propano',
    38: 'Luz',
    39: 'Basura',
    40: 'Agua y Comida',
    41: 'Impuestos Propiedad',
    42: 'Asociacion Civil',
    43: 'Gastos de Capital',
    44: 'HADIME Comida',
    45: 'HADIME Alojamiento',
    46: 'HADIME Ministerio',
    47: 'HADIME Transporte',
    48: 'HADIME Van',
    49: 'HADIME Predicadores',
    50: 'HADIME Misc',
    52: 'Gasolina Van',
    53: 'Reparaciones Van',
    54: 'Seguro Van',
}

# Solo 2026
SHEETS_2025 = []
SHEETS_2026 = ['ENE26','FEB26','MAR26','ABR26']

MONTH_MAP = {
    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12
}

def parse_sheet_date(sname):
    s = sname.upper().strip()
    mon_str = s[:3]
    yr_str  = s[3:]
    mon = MONTH_MAP.get(mon_str, 1)
    yr  = int('20' + yr_str) if len(yr_str) == 2 else int(yr_str)
    return yr, mon

def safe_float(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except Exception:
        return None

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    all_sheets = SHEETS_2025 + SHEETS_2026

    transactions = []
    cat_summary  = {}
    total_income  = 0.0
    total_expense = 0.0

    for sname in all_sheets:
        if sname not in wb.sheetnames:
            print(f'  [SKIP] Hoja no encontrada: {sname}')
            continue
        yr, mon = parse_sheet_date(sname)
        ws   = wb[sname]
        rows = list(ws.iter_rows(values_only=True))

        for row in rows[3:]:  # filas de datos (saltamos 3 cabeceras)
            if row[1] is None:
                continue
            date_ref = str(row[1]).strip()
            # Ignorar filas de totales / saldos
            skip_keywords = ['saldo', 'total', 'cerrar', 'open bal', 'balance']
            if any(kw in date_ref.lower() for kw in skip_keywords):
                continue

            # Construir fecha: usar dia del ref si es numero, si no usar dia 1
            day = 1
            # Algunos refs son "J 1", "E 2", "F 15" etc. — extraer numero
            parts = date_ref.replace('-','').split()
            for p in parts:
                try:
                    d = int(p)
                    if 1 <= d <= 31:
                        day = d
                        break
                except Exception:
                    pass
            # Validar dia
            try:
                txn_date = datetime(yr, mon, min(day, 28)).strftime('%d/%m/%Y')
            except Exception:
                txn_date = datetime(yr, mon, 1).strftime('%d/%m/%Y')

            # Ingresos
            for col_i, cat_name in INCOME_COLS.items():
                if col_i < len(row):
                    amount = safe_float(row[col_i])
                    if amount is not None and amount != 0:
                        amount = abs(amount)
                        transactions.append({
                            'transaction_date':  txn_date,
                            'transaction_type':  'income',
                            'category':          cat_name,
                            'amount':            f'{amount:.2f}',
                            'description':       f'[{sname}] {cat_name}',
                            'notes':             f'Ref: {date_ref}',
                            'payment_method':    'transferencia',
                            'reference_number':  f'{sname}-{date_ref}',
                        })
                        total_income += amount
                        cat_summary[cat_name] = cat_summary.get(cat_name, {'count': 0, 'total': 0.0, 'type': 'income'})
                        cat_summary[cat_name]['count'] += 1
                        cat_summary[cat_name]['total'] += amount

            # Egresos
            for col_i, cat_name in EXPENSE_COLS.items():
                if col_i < len(row):
                    amount = safe_float(row[col_i])
                    if amount is not None and amount != 0:
                        amount = abs(amount)
                        transactions.append({
                            'transaction_date':  txn_date,
                            'transaction_type':  'expense',
                            'category':          cat_name,
                            'amount':            f'{amount:.2f}',
                            'description':       f'[{sname}] {cat_name}',
                            'notes':             f'Ref: {date_ref}',
                            'payment_method':    'transferencia',
                            'reference_number':  f'{sname}-{date_ref}',
                        })
                        total_expense += amount
                        cat_summary[cat_name] = cat_summary.get(cat_name, {'count': 0, 'total': 0.0, 'type': 'expense'})
                        cat_summary[cat_name]['count'] += 1
                        cat_summary[cat_name]['total'] += amount

    wb.close()

    # --- Resumen en consola ---
    print(f'\n=== RESUMEN DEL ANALISIS (2026) ===')
    print(f'Hojas 2026 : {len(SHEETS_2026)} meses (ENE-ABR)')
    print(f'Transacciones    : {len(transactions)}')
    print(f'Total ingresos   : ${total_income:,.2f} MXN')
    print(f'Total egresos    : ${total_expense:,.2f} MXN')
    print(f'Balance neto     : ${total_income - total_expense:,.2f} MXN')
    print()
    print('Categorias (top 20 por monto):')
    sorted_cats = sorted(cat_summary.items(), key=lambda x: x[1]['total'], reverse=True)
    for i, (cat, data) in enumerate(sorted_cats[:20], 1):
        tipo = 'ING' if data['type'] == 'income' else 'EGR'
        print(f'  {i:2d}. [{tipo}] {cat}: {data["count"]} txns | ${data["total"]:,.0f}')

    # --- Generar CSV para importacion ---
    with open(OUT_CSV, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'transaction_date','transaction_type','category','amount',
            'description','notes','payment_method','reference_number'
        ])
        writer.writeheader()
        writer.writerows(transactions)

    print(f'\nCSV generado: {OUT_CSV}')
    print(f'Filas escritas: {len(transactions)}')

if __name__ == '__main__':
    main()
